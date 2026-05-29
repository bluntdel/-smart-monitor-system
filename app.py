from flask import Flask, render_template, request, flash, redirect, jsonify, Response, url_for, send_file
import pandas as pd
import sqlite3
import os
import json
from dotenv import load_dotenv
from datetime import date
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import threading
import uuid
import hashlib
import jwt
from functools import wraps

app = Flask(__name__)
app.secret_key = 'your-secret-key'

# JWT 配置（可以使用现有的 SECRET_KEY）
JWT_SECRET = app.secret_key
JWT_ALGORITHM = 'HS256'
JWT_EXP_DELTA_SECONDS = 7200  # 24小时
load_dotenv()

# 必须配置上传文件夹
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# SQLite 数据库文件路径（从环境变量读取，默认为 app.db）
DB_PATH = os.getenv("DB_PATH", "app.db")


def get_db():
    """返回 SQLite 数据库连接，设置 row_factory 使返回字典风格的行"""
    conn = sqlite3.connect(DB_PATH)

    conn.row_factory = sqlite3.Row  # 允许通过列名访问，如 row['column_name']
    return conn


def init_db():
    """初始化数据库表（如果不存在）"""
    conn = get_db()
    cursor = conn.cursor()
    conn.commit()
    conn.close()


# 应用启动时初始化数据库
# with app.app_context():
#     init_db()


def get_val(arr, idx):
    try:
        return str(arr[idx]).strip() if idx < len(arr) else ""
    except:
        return ""


# ---------- 独立导入接口 ----------
from datetime import datetime


def get_val(arr, idx):
    try:
        val = arr[idx]
        if val is None:
            return ""
        return str(val).strip()
    except (IndexError, TypeError):
        return ""


def is_valid(value):
    """检查值是否有效（非空、非 'None' 字符串）"""
    if value is None:
        return False
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == '' or stripped.lower() == 'none' or stripped.lower() == 'nan':
            return False
    return True


def log_import_record(cursor, filename, file_hash, batch_no, status, error_msg=None, inserted_count=0):
    """向 import_log 表插入一条记录"""
    cursor.execute(
        """INSERT INTO import_log (filename, file_hash, batch_no, status, error_msg, inserted_count)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (filename, file_hash, batch_no, status, error_msg, inserted_count)
    )


def api_response(code=200, msg="", data=None):
    return jsonify({
        "code": code,
        "msg": msg,
        "data": data if data is not None else []
    })

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        auth_header = request.headers.get('Authorization')
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
        if not token:
            return jsonify({"error": "缺少 token", "code": 403}), 200

        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user_id = payload['user_id']
            admin_flg = payload.get('admin_flg', 0)
            session_key = payload.get('session_key')
        except jwt.ExpiredSignatureError:
            return jsonify({"error": "token 已过期", "code": 403}), 200
        except Exception:
            return jsonify({"error": "无效 token", "code": 403}), 200

        # 验证 session_key（防止多设备登录导致旧 token 失效）
        if session_key is None:
            return jsonify({"error": "token 格式错误", "code": 401}), 200

        cached_key = token_cache.get(user_id)
        if not cached_key or cached_key != session_key:
            return jsonify({"error": "token 已失效，请重新登录", "code": 403}), 200

        # 可选：验证用户是否存在
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM sms_user WHERE id = ?", (user_id,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()
        if not user:
            return jsonify({"error": "用户不存在", "code": 401}), 200

        # 注入参数
        kwargs['current_user_id'] = user_id
        kwargs['current_user_admin'] = admin_flg
        return f(*args, **kwargs)

    return decorated


from functools import wraps
from flask import request, jsonify
import jwt
from app import get_db, JWT_SECRET, JWT_ALGORITHM


def model_office_permission_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        auth_header = request.headers.get('Authorization')
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
        if not token:
            return jsonify({"error": "缺少 token", "code": 403}), 200

        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user_id = payload['user_id']
            admin_flg = payload.get('admin_flg', 0)
            session_key = payload.get('session_key')
        except jwt.ExpiredSignatureError:
            return jsonify({"error": "token 已过期", "code": 403}), 200
        except Exception:
            return jsonify({"error": "无效 token", "code": 403}), 200

        # 验证 session_key 是否与缓存中一致
        cached_key = token_cache.get(user_id)
        if not cached_key or cached_key != session_key:
            return jsonify({"error": "token 已失效，请重新登录", "code": 403}), 200

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT office_id FROM sms_user WHERE id = ?", (user_id,))
        user_row = cursor.fetchone()
        if not user_row:
            cursor.close()
            conn.close()
            return jsonify({"error": "用户不存在", "code": 401}), 200

        office_id = user_row['office_id']
        # 超级管理员：admin_flg=1 且 office_id 为空（None或空串）
        is_super_admin = (admin_flg == 1) and (not office_id)
        if is_super_admin:
            allowed_models = None
        else:
            allowed_models = []
            if office_id:
                cursor.execute("SELECT model_name FROM model_office WHERE office_id = ?", (office_id,))
                rows = cursor.fetchall()
                allowed_models = [row['model_name'] for row in rows]
        cursor.close()
        conn.close()

        kwargs['allowed_models'] = allowed_models
        return f(*args, **kwargs)

    return decorated


def is_super_admin(user_row):
    """user_row 包含 admin_flg 和 office_id 字段"""
    return user_row['admin_flg'] == 1 and not user_row['office_id']

@app.route('/api/logout', methods=['GET'])
@token_required
def logout(current_user_id, current_user_admin):
    token_cache.delete(current_user_id)
    return jsonify({"code": 200, "msg": f"注销成功"}), 200

@app.route('/api/import', methods=['POST'])
@token_required
def import_excel(current_user_id, current_user_admin):
    files = request.files.getlist('file')
    if not files:
        return jsonify({"code": 400, "msg": "请至少选择一个 Excel 文件", "msg1": ""}), 200

    # 获取当前用户权限信息
    conn_auth = get_db()
    cursor_auth = conn_auth.cursor()
    cursor_auth.execute("SELECT office_id, admin_flg FROM sms_user WHERE id = ?", (current_user_id,))
    user_info = cursor_auth.fetchone()
    if not user_info:
        return jsonify({"code": 500, "msg": "用户信息不存在"}), 200
    user_office_id = user_info['office_id']
    user_admin_flg = user_info['admin_flg']
    is_super_admin = (user_admin_flg == 1) and (not user_office_id)

    allowed_models = None
    if not is_super_admin:
        allowed_models = []
        if user_office_id:
            cursor_auth.execute("SELECT model_name FROM model_office WHERE office_id = ?", (user_office_id,))
            rows = cursor_auth.fetchall()
            allowed_models = [row['model_name'] for row in rows]
    cursor_auth.close()
    conn_auth.close()

    total_inserted = 0
    success_count = 0
    duplicate_count = 0
    duplicate_files = []
    errors = []

    for file in files:
        filename = file.filename
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            errors.append(f'{filename}: 不支持的文件类型')
            continue

        safe_name = secure_filename(filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
        batch_no = None
        try:
            file.save(file_path)
            df = pd.read_excel(file_path, header=None, dtype=str)
            data = df.values.tolist()
            if len(data) < 2:
                raise ValueError('文件至少需要两行数据（表头+数据行）')

            # 计算文件数据哈希
            df_filled = df.fillna('')
            content_str = df_filled.to_csv(index=False, header=False).encode('utf-8')
            data_hash = hashlib.md5(content_str).hexdigest()
            print(f"文件 {filename} 的数据哈希值: {data_hash}")

            first_row = data[0]
            model_name = get_val(data[1], 0)

            # 1. 先校验权限
            if not is_super_admin and allowed_models is not None:
                if model_name not in allowed_models:
                    raise PermissionError(f'没有权限导入模型 {model_name}')

            # 2. 再检查重复
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("SELECT 1 FROM model_data WHERE data_hash_code = ? LIMIT 1", (data_hash,))
            if cursor.fetchone():
                duplicate_count += 1
                duplicate_files.append(filename)
                cursor.close()
                conn.close()
                continue

            jgbm_title = get_val(first_row, 1)
            jgmc_title = get_val(first_row, 2)
            sjrq_title = get_val(first_row, 3)
            field_titles = [get_val(first_row, i) for i in range(4, 24)]

            if model_name:
                # 处理 model_config
                cursor.execute("SELECT id FROM model_config WHERE model_name = ?", (model_name,))
                if not cursor.fetchone():
                    base_fields = ['model_name', 'jgbm', 'jgmc', 'sjrq'] + [f'field{i}' for i in range(1, 21)]
                    base_values = [model_name, jgbm_title, jgmc_title, sjrq_title] + field_titles
                    all_fields = []
                    all_values = []
                    for idx, field in enumerate(base_fields):
                        orig_val = base_values[idx]
                        all_fields.extend([field, f'{field}_des', f'{field}_disable'])
                        all_values.extend([orig_val, orig_val, "1"])
                    placeholders = ','.join(['?'] * len(all_fields))
                    insert_sql = f"INSERT INTO model_config ({','.join(all_fields)}) VALUES ({placeholders})"
                    cursor.execute(insert_sql, all_values)

                # 处理 model_type（默认类型4）
                cursor.execute("SELECT id FROM model_type WHERE model_name = ?", (model_name,))
                if not cursor.fetchone():
                    random_id = int(time.time() * 1000) + random.randint(0, 9999)
                    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    cursor.execute(
                        "INSERT INTO model_type (id, model_name, type_code, type_des, create_time) VALUES (?, ?, ?, ?, ?)",
                        (random_id, model_name, 4, "合规与操作风险监测模型", current_time)
                    )

            base_name = os.path.splitext(filename)[0]
            batch_no = f"{base_name}_{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}"

            # 过滤有效数据行
            valid_rows = []
            for row in data[1:]:
                jgbm_val = get_val(row, 1)
                jgmc_val = get_val(row, 2)
                sjrq_val = get_val(row, 3)
                if is_valid(jgbm_val) and is_valid(jgmc_val) and is_valid(sjrq_val):
                    valid_rows.append(row)

            if not valid_rows:
                raise ValueError('没有符合条件的数据行（机构信息为空）')

            inserted_this_file = 0
            for row in valid_rows:
                cursor.execute('''
                    INSERT INTO model_data (
                        model_name, jgbm, jgmc, sjrq, batch_no, data_hash_code,
                        field1, field2, field3, field4, field5,
                        field6, field7, field8, field9, field10,
                        field11, field12, field13, field14, field15,
                        field16, field17, field18, field19, field20
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    get_val(row, 0),
                    get_val(row, 1),
                    get_val(row, 2),
                    get_val(row, 3),
                    batch_no,
                    data_hash,
                    get_val(row, 4), get_val(row, 5), get_val(row, 6), get_val(row, 7), get_val(row, 8),
                    get_val(row, 9), get_val(row, 10), get_val(row, 11), get_val(row, 12), get_val(row, 13),
                    get_val(row, 14), get_val(row, 15), get_val(row, 16), get_val(row, 17), get_val(row, 18),
                    get_val(row, 19), get_val(row, 20), get_val(row, 21), get_val(row, 22), get_val(row, 23)
                ))
                inserted_this_file += 1

            total_inserted += inserted_this_file
            success_count += 1
            conn.commit()
            cursor.close()
            conn.close()

        except PermissionError as e:
            errors.append(f'{filename}: {str(e)}')
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception as e:
            errors.append(f'{filename}: {str(e)}')
            if os.path.exists(file_path):
                os.remove(file_path)
        finally:
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except PermissionError:
                    print(f"警告：无法删除临时文件 {file_path}")

    # 汇总日志
    if len(files) > 0:
        if success_count == len(files) and duplicate_count == 0 and not errors:
            overall_status = 'all_success'
            error_msg_summary = f"全部成功：共导入 {success_count} 个文件，合计 {total_inserted} 条记录"
        elif success_count > 0:
            overall_status = 'partial_success'
            parts = [f"成功导入 {success_count} 个文件，合计 {total_inserted} 条记录"]
            if duplicate_count > 0:
                parts.append(f"重复文件 {duplicate_count} 个（内容已存在，已跳过）：{'、'.join(duplicate_files)}")
            if errors:
                parts.append(f"失败: {'；'.join(errors)}")
            error_msg_summary = "；".join(parts)
        else:
            overall_status = 'all_failed'
            parts = []
            if duplicate_count > 0:
                parts.append(f"重复文件 {duplicate_count} 个（内容已存在，已跳过）：{'、'.join(duplicate_files)}")
            if errors:
                parts.append(f"失败: {'；'.join(errors)}")
            error_msg_summary = "；".join(parts) if parts else "所有文件均未成功导入"

        conn_log = get_db()
        cursor_log = conn_log.cursor()
        random_id = int(time.time() * 1000) + random.randint(0, 9999)
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor_log.execute(
            "INSERT INTO import_log (id, batch_no, status, error_msg, create_time) VALUES (?, ?, ?, ?, ?)",
            (random_id, batch_no, overall_status, error_msg_summary, current_time)
        )
        conn_log.commit()
        cursor_log.close()
        conn_log.close()

    # 构建返回响应（换行分隔权限错误和重复错误）
    permission_errors = []  # 权限错误（包含“没有权限”的）
    other_errors = []  # 其他非权限错误
    for err in errors:
        if '没有权限' in err:
            permission_errors.append(err)
        else:
            other_errors.append(err)

    if success_count == len(files) and duplicate_count == 0 and not errors:
        # 全部成功
        code = 200
        msg = f"成功导入 {success_count} 个文件，共插入 {total_inserted} 条记录"
        msg2 = ""
        msg1 = ""
    else:
        # 有失败或重复
        code = 200  # 部分成功或全部失败都返回200？按之前逻辑部分成功200，全部失败500。但用户未明确，暂沿用原逻辑：有成功则200，全失败500
        if success_count > 0:
            code = 200
            msg = f"成功导入 {success_count} 个文件，共插入 {total_inserted} 条记录"
        else:
            code = 200
            msg = ""
        if duplicate_count > 0:
            msg2 = f"重复文件 {duplicate_count} 个（内容已存在，已跳过）：{'、'.join(duplicate_files)}"
        else:
            msg2 = ""
        if permission_errors:
            msg1 = "\n".join(permission_errors)
        else:
            msg1 = ""
        if other_errors:
            if msg1:
                msg1 += "\n" + "\n".join(other_errors)
            else:
                msg1 = "\n".join(other_errors)

    return jsonify({
        "code": code,
        "msg": msg,
        "msg2": msg2,
        "msg1": msg1,
        "duplicate_files": duplicate_files,
        "duplicate_count": duplicate_count,
        "success_count": success_count,
        "total_inserted": total_inserted,
        "errors": errors  # 保留原始错误列表供调试
    }), 200


# ---------- 驾驶舱页面（仅 GET） ----------
@app.route('/')
def dashboard():
    """驾驶舱首页，展示统计图表"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) AS cnt FROM model_config")
    totalModel = cursor.fetchone()['cnt']

    cursor.execute("SELECT COUNT(*) AS cnt FROM model_data")
    totalData = cursor.fetchone()['cnt']

    cursor.execute("SELECT COUNT(DISTINCT model_name) AS cnt FROM model_data")
    totalOrg = cursor.fetchone()['cnt']

    today = date.today()
    cursor.execute("SELECT COUNT(*) AS cnt FROM model_data WHERE DATE(create_time)=?", (today,))
    todayImport = cursor.fetchone()['cnt']

    cursor.execute("""
        SELECT model_name, COUNT(*) AS cnt 
        FROM model_data 
        GROUP BY model_name 
        ORDER BY cnt DESC 
        LIMIT 5
    """)
    top5Models = cursor.fetchall()

    cursor.execute("""
        SELECT model_name, COUNT(*) AS count 
        FROM model_data 
        GROUP BY model_name 
        ORDER BY count DESC
    """)
    orgData = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("index.html",
                           totalModel=totalModel,
                           totalData=totalData,
                           totalOrg=totalOrg,
                           todayImport=todayImport,
                           top5Models=top5Models,
                           orgData=orgData)


@app.route('/api/chart-data-all')
@model_office_permission_required
def api_chart_data(allowed_models, **kwargs):
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')
    if not start_date or not end_date:
        return jsonify({"code": 400, "msg": f"缺少必传参数：startDate, endDate"}), 200

    if not (start_date.isdigit() and len(start_date) == 8 and end_date.isdigit() and len(end_date) == 8):
        return jsonify({"code": 400, "msg": f"日期格式必须为 yyyymmdd"}), 200

    # 权限过滤
    if allowed_models is None:
        model_where = ""
        model_params = []
    elif allowed_models:
        placeholders = ','.join(['?'] * len(allowed_models))
        model_where = f"AND md.model_name IN ({placeholders})"
        model_params = allowed_models
    else:
        return api_response(data=[])  # 无权限返回空

    try:
        conn = get_db()
        cursor = conn.cursor()
        sql_stats = f"""
            SELECT mt.type_code, COUNT(*) AS count
            FROM model_data md
            INNER JOIN model_type mt ON md.model_name = mt.model_name
            WHERE md.sjrq BETWEEN ? AND ? {model_where}
            GROUP BY mt.type_code
        """
        params = [start_date, end_date] + model_params
        cursor.execute(sql_stats, params)
        rows = cursor.fetchall()
        stats_map = {row["type_code"]: row["count"] for row in rows}

        # 类型描述（不变）
        cursor.execute(
            "SELECT DISTINCT type_code, type_des FROM model_type WHERE type_code IN (1,2,3,4,5) ORDER BY type_code")
        type_rows = cursor.fetchall()
        type_desc_map = {row["type_code"]: row["type_des"] for row in type_rows} if type_rows else {1: "类型1", 2: "类型2",
                                                                                                    3: "类型3", 4: "类型4",
                                                                                                    5: "类型5"}
        all_types = [1, 2, 3, 4, 5]
        result = [{"type_code": tc, "typeDes": type_desc_map.get(tc, f"类型{tc}"), "count": stats_map.get(tc, 0)} for tc
                  in all_types]
        return api_response(data=result)
    except Exception as e:
        print(e)

        return api_response(data=[])  # 无权限返回空


@app.route('/api/chart-org-data')
@model_office_permission_required
def api_chart_org_data(allowed_models, **kwargs):
    type_code = request.args.get('typeCode')
    month = request.args.get('month')
    jgmc = request.args.get('jgmc')
    if not type_code or not month or not jgmc:
        return jsonify({"code": 400, "msg": f"type_code、month、jgmc缺少必传参数"}), 200

    # 模型权限过滤
    if allowed_models is None:
        model_where = ""
        model_params = []
    elif allowed_models:
        placeholders = ','.join(['?'] * len(allowed_models))
        model_where = f"AND t1.model_name IN ({placeholders})"
        model_params = allowed_models
    else:
        return api_response(data=[])  # 无权限返回空

    sql = f"""
        SELECT t1.model_name, COUNT(t1.model_name) AS count
        FROM model_data t1
        LEFT JOIN model_type t2 ON t1.model_name = t2.model_name
        WHERE t2.type_code = ? AND SUBSTR(t1.sjrq, 1, 6) = ? AND t1.jgmc = ? {model_where}
        GROUP BY t1.model_name
    """
    params = [type_code, month, jgmc] + model_params
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        data = [{"model_name": row["model_name"], "count": row["count"]} for row in rows]
        cursor.close()
        conn.close()
        return api_response(data=data)  # 无权限返回空

    except Exception as e:
        print("查询错误：", e)
        return api_response(data=[])  # 无权限返回空


@app.route('/api/chart-data-monthly-all')
@model_office_permission_required
def chart_data_monthly_type(allowed_models, **kwargs):
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')
    if not start_date or not end_date:
        return jsonify({"code": 400, "msg": f"缺少必传参数：startDate, endDate"}), 200

    if not (start_date.isdigit() and len(start_date) == 8 and end_date.isdigit() and len(end_date) == 8):
        return jsonify({"code": 400, "msg": f"日期格式错误"}), 200

    # 权限过滤
    if allowed_models is None:
        model_where = ""
        model_params = []
    elif allowed_models:
        placeholders = ','.join(['?'] * len(allowed_models))
        model_where = f"AND md.model_name IN ({placeholders})"
        model_params = allowed_models
    else:
        return api_response(data=[])  # 无权限返回空

    try:
        conn = get_db()
        cursor = conn.cursor()
        # 类型描述
        cursor.execute(
            "SELECT DISTINCT type_code, type_des FROM model_type WHERE type_code IN (1,2,3,4,5) ORDER BY type_code")
        type_rows = cursor.fetchall()
        type_desc_map = {row["type_code"]: row["type_des"] for row in type_rows} if type_rows else {1: "类型1", 2: "类型2",
                                                                                                    3: "类型3", 4: "类型4",
                                                                                                    5: "类型5"}

        sql = f"""
            SELECT SUBSTR(md.sjrq, 1, 6) AS month, mt.type_code, COUNT(*) AS count
            FROM model_data md
            INNER JOIN model_type mt ON md.model_name = mt.model_name
            WHERE md.sjrq BETWEEN ? AND ? {model_where}
            GROUP BY SUBSTR(md.sjrq, 1, 6), mt.type_code
            ORDER BY month, mt.type_code
        """
        params = [start_date, end_date] + model_params
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        data_map = {(row["month"], row["type_code"]): row["count"] for row in rows}

        # 生成月份列表
        start_year, start_month = int(start_date[:4]), int(start_date[4:6])
        end_year, end_month = int(end_date[:4]), int(end_date[4:6])
        months = []
        y, m = start_year, start_month
        while (y < end_year) or (y == end_year and m <= end_month):
            months.append(f"{y}{m:02d}")
            m += 1
            if m == 13:
                m = 1
                y += 1

        all_types = [1, 2, 3, 4, 5]
        result = []
        for ym in months:
            data_list = [
                {"type_code": tc, "typeDes": type_desc_map.get(tc, f"类型{tc}"), "count": data_map.get((ym, tc), 0)} for
                tc in all_types]
            result.append({"month": f"{ym[:4]}-{ym[4:]}", "data": data_list})
        return api_response(data=result)  # 无权限返回空
    except Exception as e:
        print(e)
        return jsonify({"error": str(e)}), 500


@app.route('/api/chart-data')
@model_office_permission_required
def api_chart_data1(allowed_models, **kwargs):
    type_code = request.args.get('typeCode')
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')
    jgbm = request.args.get('jgbm')
    if not type_code or not start_date or not end_date:
        return jsonify({"error": "缺少必传参数"}), 400
    if not (start_date.isdigit() and len(start_date) == 8 and end_date.isdigit() and len(end_date) == 8):
        return jsonify({"error": "日期格式错误"}), 400

    # 权限过滤
    if allowed_models is None:
        model_where = ""
        model_params = []
    elif allowed_models:
        placeholders = ','.join(['?'] * len(allowed_models))
        model_where = f"AND t1.model_name IN ({placeholders})"
        model_params = allowed_models
    else:
        return api_response(data=[])  # 无权限返回空

    sql = f"""
        SELECT t1.model_name, COUNT(t1.model_name) AS count
        FROM model_data t1
        INNER JOIN model_type t2 ON t1.model_name = t2.model_name
        WHERE t2.type_code = ? AND t1.sjrq BETWEEN ? AND ? {model_where}
    """
    params = [type_code, start_date, end_date] + model_params
    if jgbm:
        sql += " AND t1.jgbm = ?"
        params.append(jgbm)
    sql += " GROUP BY t1.model_name"

    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        data = [{"model_name": row["model_name"], "count": row["count"]} for row in rows]
        return api_response(data=data)  # 无权限返回空
    except Exception as e:
        print(e)
        return api_response(data=[])  # 无权限返回空


@app.route('/api/chart-data-detail1')
@model_office_permission_required
def chart_data_detail1(allowed_models, **kwargs):
    model_name = request.args.get('modelName')
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')
    jgbm = request.args.get('jgbm')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    if page < 1:
        page = 1
    if per_page < 1:
        per_page = 20
    offset = (page - 1) * per_page

    if not model_name or not start_date or not end_date:
        return Response(json.dumps({"error": "缺少 modelName 或 startDate/endDate 参数", "code": 400}, ensure_ascii=False),
                        mimetype='application/json')  # 权限检查：如果非管理员且 model_name 不在允许列表中，直接返回空
    if allowed_models is not None and model_name not in allowed_models:
        return jsonify({"headers": [], "data": [], "total": 0, "page": page, "per_page": per_page})

        # 可选：校验日期格式（8位数字）
    if not (start_date.isdigit() and len(start_date) == 8 and end_date.isdigit() and len(end_date) == 8):
        return Response(json.dumps({"error": "日期格式必须为 yyyymmdd，如 20260101", "code": 400}, ensure_ascii=False),
                        mimetype='application/json')

    try:
        conn = get_db()
        cursor = conn.cursor()

        # 基础字段列表（顺序决定了展示顺序）
        base_fields = ['model_name', 'jgmc', 'jgbm', 'sjrq'] + [f'field{i}' for i in range(1, 21)]

        # 查询 model_config，获取所有 _des 和 _disable
        select_config = []
        for f in base_fields:
            select_config.append(f + '_des')
            select_config.append(f + '_disable')
        sql_config = f"SELECT {','.join(select_config)} FROM model_config WHERE model_name = ? LIMIT 1"
        cursor.execute(sql_config, (model_name,))
        config_row = cursor.fetchone()
        if not config_row:
            return Response(json.dumps({"error": "未找到模型配置", "code": 404}, ensure_ascii=False),
                            mimetype='application/json')

        # 构建标题列表和需要查询的字段列表
        headers = []
        data_fields = []
        for f in base_fields:
            disable_val = config_row[f + '_disable']
            if disable_val == "1":
                title = config_row[f + '_des'] or f
                headers.append(title)
                data_fields.append(f)

        if not data_fields:
            return Response(json.dumps({"headers": [], "data": [], "total": 0, "page": page, "per_page": per_page},
                                       ensure_ascii=False),
                            mimetype='application/json')

        # 构建 WHERE 条件（日期范围）
        where_clause = "model_name = ? AND sjrq BETWEEN ? AND ?"
        params = [model_name, start_date, end_date]
        if jgbm:
            where_clause += " AND jgbm = ?"
            params.append(jgbm)

        # 1. 查询总记录数
        count_sql = f"SELECT COUNT(*) AS total FROM model_data WHERE {where_clause}"
        cursor.execute(count_sql, params)
        total = cursor.fetchone()['total']

        # 2. 分页查询数据
        select_data = ','.join(data_fields)
        data_sql = f"""
               SELECT {select_data}
               FROM model_data
               WHERE {where_clause}
               LIMIT ? OFFSET ?
           """
        data_params = params + [per_page, offset]
        cursor.execute(data_sql, data_params)
        rows = cursor.fetchall()
        data = [list(row) for row in rows]

        cursor.close()
        conn.close()

        result = {
            "code": 200,
            "headers": headers,
            "data": data,
            "total": total,
            "page": page,
            "per_page": per_page
        }
        return Response(json.dumps(result, ensure_ascii=False), mimetype="application/json")
    except Exception as e:
        print("查询错误：", e)
        return Response(json.dumps({"error": str(e), "code": 500}, ensure_ascii=False),
                        mimetype='application/json')


@app.route('/api/export-excel')
@model_office_permission_required
def export_excel(allowed_models, **kwargs):
    model_name = request.args.get('modelName')
    month = request.args.get('month')
    jgbm = request.args.get('jgbm')
    if not model_name or not month:
        return jsonify({"code": 400, "msg": f"缺少必传参数：model_name或 month"}), 200

    if allowed_models is not None and model_name not in allowed_models:
        return jsonify({"code": 400, "msg": f"无权访问该模型"}), 200

    # 其余代码不变

    try:
        conn = get_db()
        cursor = conn.cursor()

        # 1. 基础字段顺序（决定最终 Excel 列顺序）
        base_fields = ['model_name', 'jgmc', 'jgbm', 'sjrq'] + [f'field{i}' for i in range(1, 21)]

        # 2. 查询 model_config，获取所有字段的 _des 和 _disable
        select_config = []
        for f in base_fields:
            select_config.append(f + '_des')
            select_config.append(f + '_disable')
        # 注意：不查询原始值，只取 _des, _disable
        sql_config = f"SELECT {','.join(select_config)} FROM model_config WHERE model_name = ? LIMIT 1"
        cursor.execute(sql_config, (model_name,))
        config_row = cursor.fetchone()
        if not config_row:
            return jsonify({"code": 400, "msg": f"未找到模型配置"}), 200

        # 3. 构建表头和数据字段列表
        headers = []
        data_fields = []
        for f in base_fields:
            disable_val = config_row[f + '_disable']
            if disable_val == "1":  # 只有启用状态才导出
                title = config_row[f + '_des'] or f
                headers.append(title)
                data_fields.append(f)

        if not data_fields:
            # 没有可导出字段，返回空 Excel 或提示错误
            return jsonify({"code": 400, "msg": f"无可导出的字段，请检查模型配置"}), 200

        # 4. 查询 model_data（不分页，导出所有符合条件的记录）
        where_clause = "model_name = ? AND SUBSTR(sjrq, 1, 6) = ?"
        params = [model_name, month]
        if jgbm:
            where_clause += " AND jgbm = ?"
            params.append(jgbm)

        select_data = ','.join(data_fields)
        data_sql = f"""
            SELECT {select_data}
            FROM model_data
            WHERE {where_clause}
        """
        cursor.execute(data_sql, params)
        rows = cursor.fetchall()
        data = [list(row) for row in rows]

        cursor.close()
        conn.close()

        # 5. 生成 Excel 文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{model_name}_{month}"

        # 写入表头
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 写入数据行
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 自动调整列宽
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = min(max_len + 2, 40)
            ws.column_dimensions[col_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{model_name}_{month}{'_' + jgbm if jgbm else ''}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print("导出失败：", e)
        return jsonify({"code": 500, "msg": str(e)}), 200


@app.route('/api/org-list')
@model_office_permission_required
def org_list(allowed_models, **kwargs):
    model_name = request.args.get('modelName')
    # 构建权限条件
    if allowed_models is None:
        model_where = ""
        model_params = []
    elif allowed_models:
        placeholders = ','.join(['?'] * len(allowed_models))
        model_where = f"AND model_name IN ({placeholders})"
        model_params = allowed_models
    else:
        return api_response(data=[])  # 无权限返回空

    try:
        conn = get_db()
        cursor = conn.cursor()
        if model_name:
            # 额外增加 model_name 筛选，同时仍要满足权限条件
            if allowed_models is not None and model_name not in allowed_models:
                return api_response(data=[])  # 无权限返回空
            sql = f"""
                SELECT DISTINCT jgmc, jgbm
                FROM model_data
                WHERE model_name = ? AND jgmc IS NOT NULL AND jgbm IS NOT NULL {model_where}
                ORDER BY jgmc
            """
            params = [model_name] + model_params
        else:
            sql = f"""
                SELECT DISTINCT jgmc, jgbm
                FROM model_data
                WHERE jgmc IS NOT NULL AND jgbm IS NOT NULL {model_where}
                ORDER BY jgmc
            """
            params = model_params
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        result = [{"jgmc": row["jgmc"], "jgbm": row["jgbm"]} for row in rows]
        return api_response(data=result)  # 无权限返回空
    except Exception as e:
        return jsonify({"code": 400, "msg": str(e)}), 200


@app.route('/api/model-list')
@model_office_permission_required
def model_list(allowed_models, **kwargs):
    type_code = request.args.get('typeCode')
    try:
        conn = get_db()
        cursor = conn.cursor()

        if allowed_models is None:
            # 超级管理员：所有模型
            sql = """
                SELECT DISTINCT mc.model_name
                FROM model_config mc
                LEFT JOIN model_type mt ON mc.model_name = mt.model_name
                WHERE mc.model_name IS NOT NULL AND mc.model_name != ''
            """
            params = []
            if type_code:
                sql += " AND mt.type_code = ?"
                params.append(type_code)
            sql += " ORDER BY mc.model_name"
        elif allowed_models:
            # 非管理员：仅允许列表中的模型
            placeholders = ','.join(['?'] * len(allowed_models))
            sql = f"""
                SELECT DISTINCT mc.model_name
                FROM model_config mc
                LEFT JOIN model_type mt ON mc.model_name = mt.model_name
                WHERE mc.model_name IN ({placeholders}) AND mc.model_name IS NOT NULL AND mc.model_name != ''
            """
            params = allowed_models.copy()
            if type_code:
                sql += " AND mt.type_code = ?"
                params.append(type_code)
            sql += " ORDER BY mc.model_name"
        else:
            return api_response(data=[])  # 无权限返回空

        cursor.execute(sql, params)
        rows = cursor.fetchall()
        result = [{"model_name": row["model_name"]} for row in rows]
        cursor.close()
        conn.close()
        return api_response(data=result)  # 无权限返回空
    except Exception as e:
        print("查询模型列表错误：", e)
        return jsonify({"code": 500, "msg": str(e)}), 500

@app.route('/api/model-listAll')
@model_office_permission_required
def model_list_all(allowed_models, **kwargs):

    try:
        sql = f"""
            SELECT DISTINCT model_name
            FROM model_type
            WHERE model_name IS NOT NULL AND model_name != ''
        """
        sql += " ORDER BY create_time"

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
        # 格式化返回结果
        result = [{"model_name": row["model_name"]} for row in rows]
        return api_response(data=result)  # 无权限返回空
    except Exception as e:
        return jsonify({"code": 400, "msg": str(e)}), 500

@app.route('/api/org-type-stats')
@model_office_permission_required
def org_type_stats(allowed_models, **kwargs):
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')
    if not start_date or not end_date:
        return jsonify({"code": 400, "msg": f"缺少参数"}), 200

    jgmc = request.args.get('jgmc')
    if not (start_date.isdigit() and len(start_date) == 8 and end_date.isdigit() and len(end_date) == 8):
        return jsonify({"code": 400, "msg": f"日期格式错误"}), 200

    if allowed_models is None:
        model_where = ""
        model_params = []
    elif allowed_models:
        placeholders = ','.join(['?'] * len(allowed_models))
        model_where = f"AND md.model_name IN ({placeholders})"
        model_params = allowed_models
    else:
        return api_response(data=[])  # 无权限返回空

    try:
        conn = get_db()
        cursor = conn.cursor()
        sql = f"""
            SELECT 
                md.jgmc AS 机构名称,
                md.jgbm AS 机构编码,
                COUNT(CASE WHEN mt.type_code = 1 THEN 1 END) AS 模型类型1记录数,
                COUNT(CASE WHEN mt.type_code = 2 THEN 1 END) AS 模型类型2记录数,
                COUNT(CASE WHEN mt.type_code = 3 THEN 1 END) AS 模型类型3记录数,
                COUNT(CASE WHEN mt.type_code = 4 THEN 1 END) AS 模型类型4记录数,
                COUNT(CASE WHEN mt.type_code = 5 THEN 1 END) AS 模型类型5记录数
            FROM model_data md
            LEFT JOIN model_type mt ON md.model_name = mt.model_name
            WHERE md.sjrq BETWEEN ? AND ? {model_where}
        """
        params = [start_date, end_date] + model_params
        if jgmc:
            sql += " AND md.jgmc = ?"
            params.append(jgmc)
        sql += " GROUP BY md.jgmc, md.jgbm ORDER BY md.jgmc"
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        result = [dict(row) for row in rows]
        return api_response(data=result)  # 无权限返回空
    except Exception as e:
        return jsonify({"code": 400, "msg": str(e)}), 500


@app.route('/api/org-type-stats2')
@model_office_permission_required
def org_type_stats2(allowed_models, **kwargs):
    # 必传参数：startDate, endDate
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')

    # 校验必传参数
    if not start_date or not end_date:
        return jsonify({"code": 400, "msg": f"缺少必传参数：startDate 和 endDate"}), 200

    # 可选参数：jgmc（机构名称）
    jgmc = request.args.get('jgmc')

    if allowed_models is None:
        model_where = ""
        model_params = []
    elif allowed_models:
        placeholders = ','.join(['?'] * len(allowed_models))
        model_where = f"AND md.model_name IN ({placeholders})"
        model_params = allowed_models
    else:
        return api_response(data=[])  # 无权限返回空

    try:
        conn = get_db()
        cursor = conn.cursor()

        # 基础 SQL：统计每个机构的各模型类型记录数，限定日期范围
        sql = f"""
            SELECT 
                md.jgmc AS 机构名称,
                md.jgbm AS 机构编码,
                COUNT(CASE WHEN mt.type_code = 1 THEN 1 END) AS 模型类型1记录数,
                COUNT(CASE WHEN mt.type_code = 2 THEN 1 END) AS 模型类型2记录数,
                COUNT(CASE WHEN mt.type_code = 3 THEN 1 END) AS 模型类型3记录数,
                COUNT(CASE WHEN mt.type_code = 4 THEN 1 END) AS 模型类型4记录数,
                COUNT(CASE WHEN mt.type_code = 5 THEN 1 END) AS 模型类型5记录数
            FROM model_data md
            LEFT JOIN model_type mt ON md.model_name = mt.model_name
            WHERE md.sjrq BETWEEN ? AND ? {model_where}
        """
        params = [start_date, end_date] + model_params

        # 可选：按机构名称筛选
        if jgmc:
            sql += " AND md.jgmc = ?"
            params.append(jgmc)

        sql += " GROUP BY md.jgmc, md.jgbm ORDER BY md.jgmc"

        cursor.execute(sql, params)
        rows = cursor.fetchall()

        # 将 sqlite3.Row 对象转换为字典列表
        result = [dict(row) for row in rows]

        cursor.close()
        conn.close()
        return api_response(data=result)  # 无权限返回空

    except Exception as e:
        print("机构类型统计查询错误：", e)
        return jsonify({"code": 400, "msg": str(e)}), 500


@app.route('/api/chart-data-monthly')
@model_office_permission_required
def chart_data_monthly(allowed_models, **kwargs):
    type_code = request.args.get('typeCode')
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')
    jgbm = request.args.get('jgbm')
    if not type_code or not start_date or not end_date:
        return jsonify({"code": 400, "msg": f"缺少参数type_code、start_date、end_date"}), 200
    if not (start_date.isdigit() and len(start_date) == 8 and end_date.isdigit() and len(end_date) == 8):
        return jsonify({"code": 400, "msg": f"日期格式错误"}), 200

    # 获取有权限的模型列表
    if allowed_models is None:
        # 管理员：所有模型（需从数据中获取去重列表）
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT model_name FROM model_data ORDER BY model_name")
        model_rows = cursor.fetchall()
        allowed_models = [row['model_name'] for row in model_rows]
        cursor.close()
        conn.close()
    elif not allowed_models:
        return api_response(data=[])  # 无权限返回空

    placeholders = ','.join(['?'] * len(allowed_models))
    try:
        conn = get_db()
        cursor = conn.cursor()
        sql = f"""
            SELECT 
                SUBSTR(md.sjrq, 1, 6) AS month,
                md.model_name,
                COUNT(*) AS count
            FROM model_data md
            INNER JOIN model_type mt ON md.model_name = mt.model_name
            WHERE mt.type_code = ?
              AND md.sjrq BETWEEN ? AND ?
              AND md.model_name IN ({placeholders})
        """
        params = [type_code, start_date, end_date] + allowed_models
        if jgbm:
            sql += " AND md.jgbm = ?"
            params.append(jgbm)
        sql += " GROUP BY SUBSTR(md.sjrq, 1, 6), md.model_name ORDER BY month, md.model_name"
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        data_map = {(row["month"], row["model_name"]): row["count"] for row in rows}

        # 生成月份列表
        start_year, start_month = int(start_date[:4]), int(start_date[4:6])
        end_year, end_month = int(end_date[:4]), int(end_date[4:6])
        months = []
        y, m = start_year, start_month
        while (y < end_year) or (y == end_year and m <= end_month):
            months.append(f"{y}{m:02d}")
            m += 1
            if m == 13:
                m = 1
                y += 1

        result = []
        for ym in months:
            model_counts = []
            for mn in allowed_models:
                count = data_map.get((ym, mn), 0)
                model_counts.append({"model_name": mn, "count": count})
            result.append({"month": f"{ym[:4]}-{ym[4:]}", "data": model_counts})
        return api_response(data=result)  # 无权限返回空
    except Exception as e:
        print(e)
        return jsonify({"code": 400, "msg": str(e)}), 500


@app.route('/api/model-config')
@token_required  # 需要登录
def get_model_config(current_user_id, current_user_admin):
    """
    根据模型名称获取 model_config 完整配置信息，并关联 model_type 获取类型代码
    返回格式：{"id": 记录ID, "type_code": 类型代码, "data": [{"字段名": 原值, "字段名_des": 描述, "字段名_disable": 禁用标志}, ...]}
    """
    model_name = request.args.get('model_name')
    if not model_name:
        return jsonify({"code": 400, "msg": f"缺少 model_name 参数"}), 200

    try:
        conn = get_db()
        cursor = conn.cursor()
        # 获取 model_config 所有列名
        cursor.execute("PRAGMA table_info(model_config)")
        all_columns = [row['name'] for row in cursor.fetchall()]
        sql = f"SELECT {','.join(all_columns)} FROM model_config WHERE model_name = ?"
        cursor.execute(sql, (model_name,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"code": 400, "msg": f"未找到模型: {model_name}"}), 200

        row_dict = dict(row)
        record_id = row_dict.get('id')

        # 关联查询 model_type 中的 type_code
        cursor.execute("SELECT type_code FROM model_type WHERE model_name = ? LIMIT 1", (model_name,))
        type_row = cursor.fetchone()
        type_code = type_row['type_code'] if type_row else None

        # 基础字段列表（决定返回顺序）
        base_fields = ['model_name', 'jgmc', 'jgbm', 'sjrq'] + [f'field{i}' for i in range(1, 21)]

        data = []
        for field in base_fields:
            original = row_dict.get(field, "")
            des = row_dict.get(f"{field}_des", "")
            disable = row_dict.get(f"{field}_disable", "")
            item = {
                field: original,
                f"{field}_des": des,
                f"{field}_disable": disable
            }
            data.append(item)

        result = {
            "id": record_id,
            "type_code": type_code,  # 新增字段，与 id 平级
            "data": data,
            "code":200
        }

        cursor.close()
        conn.close()
        return result  # 无权限返回空

    except Exception as e:
        print("查询模型配置错误：", e)
        return jsonify({"code": 400, "msg": str(e)}), 500


import time
import random


@app.route('/api/model-config-upd', methods=['PUT', 'POST', 'GET'])
@token_required  # 需要登录
def update_model_config(current_user_id, current_user_admin):
    data = request.get_json()
    if not data:
        return jsonify({"code": 400, "msg": f"请求体不能为空"}), 200

    required_fields = ['id', 'type_code', 'type_des']
    for field in required_fields:
        if field not in data:
            return jsonify({"code": 400, "msg": f"缺少必传字段: {field}"}), 200

    record_id = data['id']
    type_code = data['type_code']
    type_des = data['type_des']

    try:
        conn = get_db()
        cursor = conn.cursor()

        # 获取 model_name
        cursor.execute("SELECT model_name FROM model_config WHERE id = ?", (record_id,))
        config_row = cursor.fetchone()
        if not config_row:
            return jsonify({"code": 400, "msg": f"未找到 id 为 {record_id} 的记录"}), 200

        model_name = config_row['model_name']

        # 更新 model_config（动态字段，与原逻辑一致）
        cursor.execute("PRAGMA table_info(model_config)")
        columns = [row['name'] for row in cursor.fetchall() if row['name'] != 'id']
        update_fields = []
        params = []
        for col in columns:
            if col in data and col not in ['type_code', 'type_des']:
                update_fields.append(f"{col} = ?")
                params.append(data[col])
        if update_fields:
            params.append(record_id)
            sql = f"UPDATE model_config SET {', '.join(update_fields)} WHERE id = ?"
            cursor.execute(sql, params)

        # 处理 model_type 表
        cursor.execute("SELECT id FROM model_type WHERE model_name = ?", (model_name,))
        existing = cursor.fetchone()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if existing:
            # 更新
            cursor.execute(
                "UPDATE model_type SET type_code = ?, type_des = ? WHERE model_name = ?",
                (type_code, type_des, model_name)
            )
        else:
            # 新增：生成随机ID（时间戳毫秒 + 随机数）
            random_id = int(time.time() * 1000) + random.randint(0, 9999)
            cursor.execute(
                "INSERT INTO model_type (id, model_name, type_code, type_des, create_time) VALUES (?, ?, ?, ?, ?)",
                (random_id, model_name, type_code, type_des, current_time)
            )

        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"code": 200, "msg": f"更新成功"}), 200
    except Exception as e:
        print("更新模型配置错误：", e)
        return jsonify({"code": 400, "msg": str(e)}), 500


@app.route('/api/batch-stats')
@token_required
def batch_stats(current_user_id, current_user_admin):
    # 获取当前用户的 office_id 和 admin_flg
    conn_auth = get_db()
    cursor_auth = conn_auth.cursor()
    cursor_auth.execute("SELECT office_id, admin_flg FROM sms_user WHERE id = ?", (current_user_id,))
    user = cursor_auth.fetchone()
    cursor_auth.close()
    conn_auth.close()
    if not user:
        return jsonify({"code": 500, "msg": "用户信息不存在"}), 200

    is_super_admin = (user['admin_flg'] == 1) and (not user['office_id'])
    allowed_models = None
    if not is_super_admin:
        allowed_models = []
        if user['office_id']:
            conn_model = get_db()
            cursor_model = conn_model.cursor()
            cursor_model.execute("SELECT model_name FROM model_office WHERE office_id = ?", (user['office_id'],))
            rows = cursor_model.fetchall()
            allowed_models = [row['model_name'] for row in rows]
            cursor_model.close()
            conn_model.close()
        # 如果 allowed_models 为空且不是超级管理员，则用户没有权限查看任何模型，直接返回空数组
        if not allowed_models:
            return api_response(data=[])  # 无权限返回空

    try:
        conn = get_db()
        cursor = conn.cursor()

        order_field = request.args.get('order_field', type=str).lower()
        allowed_fields = ['model_name', 'jgmc', 'count', 'create_time']
        if not order_field or order_field not in allowed_fields:
            return jsonify({"code": 400, "msg": f"缺少必传参数 order_field 或值无效，允许值: {', '.join(allowed_fields)}"}), 200

        order_type = request.args.get('order_type', type=str).lower()
        if not order_type or order_type not in ['desc', 'asc']:
            return jsonify({"code": 400, "msg": "缺少必传参数 order_type 或值无效，允许值: desc, asc"}), 200

        limit = request.args.get('limit', default=None, type=int)
        model_name_filter = request.args.get('model_name', type=str)
        jgbm_filter = request.args.get('jgbm', type=str)

        # 构建 WHERE 子句
        where_clause = "batch_no IS NOT NULL AND batch_no != ''"
        params = []

        # 模型权限过滤
        if not is_super_admin and allowed_models:
            placeholders = ','.join(['?'] * len(allowed_models))
            where_clause += f" AND model_name IN ({placeholders})"
            params.extend(allowed_models)

        if model_name_filter:
            # 如果用户有权限过滤，且传入的 model_name 不在允许列表中，可以提前返回空（可选）
            # 这里仍然添加条件，最终结果可能为空
            where_clause += " AND model_name = ?"
            params.append(model_name_filter)
        if jgbm_filter:
            where_clause += " AND jgbm = ?"
            params.append(jgbm_filter)

        sort_dir = 'DESC' if order_type == 'desc' else 'ASC'
        field_map = {
            'model_name': 'model_name',
            'jgmc': 'jgmc',
            'count': 'COUNT(*)',
            'create_time': 'MAX(create_time)'
        }
        order_sql = field_map[order_field]

        sql = f"""
            SELECT batch_no, 
                   model_name, 
                   jgmc,
                   MAX(create_time) AS create_time,
                   COUNT(*) AS count
            FROM model_data
            WHERE {where_clause}
            GROUP BY batch_no, model_name, jgmc
            ORDER BY {order_sql} {sort_dir}
        """
        if limit is not None:
            sql += " LIMIT ?"
            params.append(limit)

        cursor.execute(sql, params)
        rows = cursor.fetchall()
        result = [
            {
                "batch_no": row["batch_no"],
                "model_name": row["model_name"],
                "jgmc": row["jgmc"],
                "create_time": row["create_time"],
                "count": row["count"]
            }
            for row in rows
        ]
        cursor.close()
        conn.close()
        return api_response(data=result)  # 无权限返回空

    except Exception as e:
        print("批次统计查询错误：", e)
        return jsonify({"code": 500, "msg": str(e)}), 500

@app.route('/api/batch-delete', methods=['POST'])
@token_required  # 需要登录
def batch_delete(current_user_id, current_user_admin):
    """
    根据批次号（支持逗号分隔多个）批量删除 model_data 中的记录
    参数：batch_nos - 逗号分隔的批次号字符串，例如 "批次1,批次2,批次3"
         （可从查询参数或 JSON 请求体中获取）
    返回：{"message": "删除成功", "deleted_count": n, "failed": [...]}
    """
    # 获取参数：优先从 JSON 获取，其次从查询参数获取
    data = request.get_json(silent=True)
    batch_nos_str = None
    if data and 'batch_nos' in data:
        batch_nos_str = data.get('batch_nos')
    else:
        batch_nos_str = request.args.get('batch_nos')

    if not batch_nos_str:
        return jsonify({"code": 400, "msg": f"缺少 batch_nos 参数"}), 200

    # 按逗号分割，去除空格和空字符串
    batch_nos = [b.strip() for b in batch_nos_str.split(',') if b.strip()]
    if not batch_nos:
        return jsonify({"code": 400, "msg": f"batch_nos 参数无效，未提供有效的批次号"}), 200

    try:
        conn = get_db()
        cursor = conn.cursor()
        total_deleted = 0
        failed = []
        for batch_no in batch_nos:
            cursor.execute("DELETE FROM model_data WHERE batch_no = ?", (batch_no,))
            deleted = cursor.rowcount
            if deleted == 0:
                failed.append({"batch_no": batch_no, "reason": "批次不存在或无数据"})
            else:
                total_deleted += deleted
        conn.commit()
        cursor.close()
        conn.close()

        if total_deleted == 0 and failed:
            return jsonify({"code": 400, "msg": "未删除任何记录", "deleted_count": 0, "failed": failed}), 200

        else:
            msg = f"成功删除 {total_deleted} 条记录"
            if failed:
                msg += f"，{len(failed)} 个批次未找到数据"
            return jsonify({"code": 200, "msg": "删除成功", "deleted_count": total_deleted, "failed": failed}), 200

    except Exception as e:
        print("批次删除错误：", e)
        return jsonify({"error": str(e)}), 500


from werkzeug.security import check_password_hash

import threading
from datetime import datetime, timedelta


class SimpleCache:
    def __init__(self):
        self._data = {}
        self._lock = threading.Lock()

    def set(self, key, value, ttl_seconds=7200):
        expiry = datetime.utcnow() + timedelta(seconds=ttl_seconds)
        with self._lock:
            self._data[key] = (value, expiry)

    def get(self, key):
        with self._lock:
            item = self._data.get(key)
            if item:
                value, expiry = item
                if datetime.utcnow() < expiry:
                    return value
                else:
                    del self._data[key]
            return None

    def delete(self, key):
        with self._lock:
            if key in self._data:
                del self._data[key]


token_cache = SimpleCache()

import uuid


@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    userid = data.get('userid')
    password_hash = data.get('password')
    if not userid or not password_hash:
        return jsonify({"code": 400, "msg": "userid和密码不能为空"}), 200

    conn = get_db()
    cursor = conn.cursor()
    # 关联 sms_office 表获取机构名称
    cursor.execute("""
        SELECT u.id, u.username, u.userid, u.password, u.admin_flg, u.office_id, o.office_name
        FROM sms_user u
        LEFT JOIN sms_office o ON u.office_id = o.office_id
        WHERE u.userid = ?
    """, (userid,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()

    if not user or user['password'] != password_hash:
        return jsonify({"code": 401, "msg": "用户名或密码错误"}), 200

    # 生成唯一会话标识
    session_key = str(uuid.uuid4())
    # 存入缓存，有效期与 JWT 过期时间相同
    token_cache.set(user['id'], session_key, ttl_seconds=JWT_EXP_DELTA_SECONDS)

    # JWT payload 中包含 session_key
    token_payload = {
        'user_id': user['id'],
        'username': user['username'],
        'userid': user['userid'],
        'admin_flg': user['admin_flg'],
        'session_key': session_key,
        'exp': datetime.utcnow() + timedelta(seconds=JWT_EXP_DELTA_SECONDS)
    }
    token = jwt.encode(token_payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

    return jsonify({
        "code": 200,
        "msg": "登录成功",
        "token": token,
        "user": {
            "id": user['id'],
            "username": user['username'],
            "userid": user['userid'],
            "admin_flg": user['admin_flg'],
            "office_id": user['office_id'],
            "office_name": user['office_name'] if user['office_name'] else ""
        }
    }), 200


@app.route('/api/user/list', methods=['GET'])
@token_required
def user_list(current_user_id, current_user_admin):
    # 获取当前用户信息
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT admin_flg, office_id FROM sms_user WHERE id = ?", (current_user_id,))
    current_user = cursor.fetchone()
    if not current_user:
        return jsonify({"code": 401, "msg": "用户不存在"}), 200

    cursor.close()
    conn.close()

    is_super = current_user['admin_flg'] == 1 and not current_user['office_id']
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    keyword = request.args.get('keyword', '').strip()
    offset = (page - 1) * per_page

    conn = get_db()
    cursor = conn.cursor()
    base_sql = """
        FROM sms_user u
        LEFT JOIN sms_office o ON u.office_id = o.office_id
        WHERE 1=1
    """
    params = []

    # 非超级管理员时，只能查看同部门用户
    if not is_super:
        if not current_user['office_id']:
            # 无部门且非超级管理员，无权查看任何用户（或返回空）
            return jsonify({"code":200,"total": 0, "page": page, "per_page": per_page, "items": []})
        base_sql += " AND u.office_id = ?"
        params.append(current_user['office_id'])

    if keyword:
        base_sql += " AND (u.username LIKE ? OR u.userid LIKE ? OR o.office_name LIKE ?)"
        like = f"%{keyword}%"
        params.extend([like, like, like])

    count_sql = f"SELECT COUNT(*) AS total {base_sql}"
    cursor.execute(count_sql, params)
    total = cursor.fetchone()['total']

    data_sql = f"""
        SELECT u.id, u.username, u.userid, u.admin_flg, u.create_time,
               u.office_id, o.office_name
        {base_sql}
        ORDER BY u.id DESC
        LIMIT ? OFFSET ?
    """
    cursor.execute(data_sql, params + [per_page, offset])
    rows = cursor.fetchall()
    items = [dict(row) for row in rows]
    cursor.close()
    conn.close()

    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "data": items,
        "code": 200
    })
# from werkzeug.security import generate_password_hash
import hashlib


@app.route('/api/user/reset-password/<int:user_id>', methods=['POST'])
@token_required  # 需要登录
def reset_password(current_user_id, current_user_admin, user_id):
    default_password = "sms12345"
    # new_hash = generate_password_hash(default_password)
    # new_hash=hashlib.md5('Jjcbrc@2026'.encode()).hexdigest()
    new_hash = "9633a63c1f0bd6ad4fdae996198539b3"
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE sms_user SET password = ? WHERE id = ?", (new_hash, user_id))
    if cursor.rowcount == 0:
        return jsonify({"code": 401, "msg": "用户不存在"}), 200
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"code": 200, "msg": f"密码已重置!"}), 200


@app.route('/api/user/add', methods=['POST'])
@token_required  # 需要登录
def add_user(current_user_id, current_user_admin):
    data = request.get_json()
    required = ['username', 'userid', 'office_id']
    for field in required:
        if field not in data:
            return jsonify({"code": 401, "msg": f"缺少字段: {field}"}), 200

    username = data['username'].strip()
    userid = data['userid'].strip()
    office_id = data['office_id']
    admin_flg = data.get('admin_flg', 0)

    # 默认密码
    # default_password = "sms12345"
    # password_hash = generate_password_hash(default_password)
    # password_hash=hashlib.md5('Jjcbrc@2026'.encode()).hexdigest()
    password_hash = "9633a63c1f0bd6ad4fdae996198539b3"

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO sms_user (username, userid, password, admin_flg, office_id) VALUES (?, ?, ?, ?, ?)",
            (username, userid, password_hash, admin_flg, office_id)
        )
        conn.commit()
        new_id = cursor.lastrowid
        cursor.close()
        conn.close()
        return jsonify({"message": "用户添加成功", "user_id": new_id}), 200
    except sqlite3.IntegrityError as e:
        return jsonify({"code": 401, "msg": "用户名或工号已存在"}), 200


@app.route('/api/user/upd', methods=['POST'])
@token_required  # 需要登录
def update_user(current_user_id, current_user_admin):
    data = request.get_json()
    if not data:
        return jsonify({"code": 401, "msg": f"请求体不能为空"}), 200

    user_id = data.get('id')
    if not user_id:
        return jsonify({"code": 401, "msg": f"缺少用户id"}), 200

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM sms_user WHERE id = ?", (user_id,))
    if not cursor.fetchone():
        return jsonify({"code": 401, "msg": f"用户不存在"}), 200

    update_fields = []
    params = []
    if 'username' in data:
        update_fields.append("username = ?")
        params.append(data['username'].strip())
    if 'userid' in data:
        update_fields.append("userid = ?")
        params.append(data['userid'].strip())
    if 'office_id' in data:
        update_fields.append("office_id = ?")
        params.append(data['office_id'])
    if 'admin_flg' in data:
        update_fields.append("admin_flg = ?")
        params.append(data['admin_flg'])
    if 'password' in data and data['password']:
        # new_hash = generate_password_hash(data['password'])
        update_fields.append("password = ?")
        params.append(data['password'])

    if not update_fields:
        return jsonify({"code": 401, "msg": f"没有需要更新的字段"}), 200

    params.append(user_id)
    sql = f"UPDATE sms_user SET {', '.join(update_fields)} WHERE id = ?"
    cursor.execute(sql, params)
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"code": 200, "msg": f"用户信息更新成功"}), 200


@app.route('/api/user/del/<int:user_id>', methods=['POST'])
@token_required  # 需要登录
def delete_user(current_user_id, current_user_admin, user_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM sms_user WHERE id = ?", (user_id,))
    if cursor.rowcount == 0:
        return jsonify({"code": 401, "msg": f"用户不存在"}), 200

    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"code": 401, "msg": f"用户已删除"}), 200


@app.route('/api/office/list', methods=['GET'])
@token_required  # 需要登录
def office_list(current_user_id, current_user_admin):
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 100, type=int)
    keyword = request.args.get('keyword', '').strip()
    offset = (page - 1) * per_page

    conn = get_db()
    cursor = conn.cursor()
    base_sql = "FROM sms_office WHERE 1=1"
    params = []
    if keyword:
        base_sql += " AND (office_code LIKE ? OR office_name LIKE ?)"
        like = f"%{keyword}%"
        params.extend([like, like])

    count_sql = f"SELECT COUNT(*) AS total {base_sql}"
    cursor.execute(count_sql, params)
    total = cursor.fetchone()['total']

    data_sql = f"""
        SELECT office_id, office_code, office_name, create_time
        {base_sql}
        ORDER BY office_id DESC
        LIMIT ? OFFSET ?
    """
    cursor.execute(data_sql, params + [per_page, offset])
    rows = cursor.fetchall()
    items = [dict(row) for row in rows]
    cursor.close()
    conn.close()
    return jsonify({
        "Code": 200,
        "total": total,
        "page": page,
        "per_page": per_page,
        "items": items
    })

# @app.route('/api/office/add', methods=['POST'])
# @token_required
# def office_add(current_user_id, current_user_admin):
#     data = request.get_json()
#     office_code = data.get('office_code', '').strip()
#     office_name = data.get('office_name', '').strip()
#     model_names = data.get('model_names', [])
#
#     if not office_code or not office_name:
#         return jsonify({"code": 400, "msg": "机构编码和名称不能为空"}), 200
#     if not isinstance(model_names, list):
#         return jsonify({"code": 400, "msg": "model_names 必须是数组"}), 200
#
#     conn = get_db()
#     cursor = conn.cursor()
#     try:
#         # 1. 插入机构
#         cursor.execute(
#             "INSERT INTO sms_office (office_code, office_name) VALUES (?, ?)",
#             (office_code, office_name)
#         )
#         office_id = cursor.lastrowid
#         conn.commit()
#
#         # 2. 关联模型
#         inserted = 0
#         for mn in set(model_names):
#             if mn and mn.strip():
#                 cursor.execute(
#                     "INSERT INTO model_office (office_id, model_name) VALUES (?, ?)",
#                     (office_id, mn.strip())
#                 )
#                 inserted += 1
#         conn.commit()
#         cursor.close()
#         conn.close()
#         return jsonify({
#             "code": 200,
#             "msg": f"机构添加成功，并关联 {inserted} 个模型",
#             "office_id": office_id
#         }), 200
#
#     except sqlite3.IntegrityError as e:
#         return jsonify({"code": 400, "msg": "机构编码或名称已存在"}), 200
#     except Exception as e:
#         return jsonify({"code": 500, "msg": str(e)}), 200



@app.route('/api/office/new', methods=['POST'])
@token_required
def office_new(current_user_id, current_user_admin):
    data = request.get_json()
    office_id = data.get('office_id')
    office_code = data.get('office_code', '').strip()
    office_name = data.get('office_name', '').strip()
    model_names = data.get('model_names', [])

    if not isinstance(model_names, list):
        return jsonify({"code": 400, "msg": "model_names 必须是数组"}), 200

    conn = get_db()
    cursor = conn.cursor()
    try:
        if office_id is not None:
            # 更新模式：检查机构是否存在
            cursor.execute("SELECT office_id FROM sms_office WHERE office_id = ?", (office_id,))
            if not cursor.fetchone():
                return jsonify({"code": 400, "msg": "机构不存在"}), 200

            # 唯一性校验：如果修改了 office_code，检查是否与其他机构重复
            if office_code:
                cursor.execute("SELECT office_id FROM sms_office WHERE office_code = ? AND office_id != ?", (office_code, office_id))
                if cursor.fetchone():
                    return jsonify({"code": 400, "msg": f"机构编码 '{office_code}' 已存在"}), 200
            # 唯一性校验：如果修改了 office_name，检查是否与其他机构重复
            if office_name:
                cursor.execute("SELECT office_id FROM sms_office WHERE office_name = ? AND office_id != ?", (office_name, office_id))
                if cursor.fetchone():
                    return jsonify({"code": 400, "msg": f"机构名称 '{office_name}' 已存在"}), 200

            # 更新机构基本信息
            update_fields = []
            params = []
            if office_code:
                update_fields.append("office_code = ?")
                params.append(office_code)
            if office_name:
                update_fields.append("office_name = ?")
                params.append(office_name)
            if update_fields:
                sql = f"UPDATE sms_office SET {', '.join(update_fields)} WHERE office_id = ?"
                params.append(office_id)
                cursor.execute(sql, params)

            # 更新模型关联（全量替换）
            cursor.execute("DELETE FROM model_office WHERE office_id = ?", (office_id,))
            inserted = 0
            for mn in set(model_names):
                if mn and mn.strip():
                    cursor.execute(
                        "INSERT INTO model_office (office_id, model_name) VALUES (?, ?)",
                        (office_id, mn.strip())
                    )
                    inserted += 1
            conn.commit()
            msg = f"成功更新机构信息及模型关联，共关联 {inserted} 个模型"
            result_office_id = office_id
        else:
            # 新增模式：必须提供编码和名称
            if not office_code or not office_name:
                return jsonify({"code": 400, "msg": "新增机构时必须提供 office_code 和 office_name"}), 200
            # 唯一性校验（新增时无需排除自身）
            cursor.execute("SELECT office_id FROM sms_office WHERE office_code = ?", (office_code,))
            if cursor.fetchone():
                return jsonify({"code": 400, "msg": f"机构编码 '{office_code}' 已存在"}), 200
            cursor.execute("SELECT office_id FROM sms_office WHERE office_name = ?", (office_name,))
            if cursor.fetchone():
                return jsonify({"code": 400, "msg": f"机构名称 '{office_name}' 已存在"}), 200

            cursor.execute(
                "INSERT INTO sms_office (office_code, office_name) VALUES (?, ?)",
                (office_code, office_name)
            )
            new_id = cursor.lastrowid
            inserted = 0
            for mn in set(model_names):
                if mn and mn.strip():
                    cursor.execute(
                        "INSERT INTO model_office (office_id, model_name) VALUES (?, ?)",
                        (new_id, mn.strip())
                    )
                    inserted += 1
            conn.commit()
            msg = f"机构添加成功，并关联 {inserted} 个模型"
            result_office_id = new_id

        cursor.close()
        conn.close()
        return jsonify({"code": 200, "msg": msg, "office_id": result_office_id}), 200
    except sqlite3.IntegrityError as e:
        # 兜底处理，通常不会走到这里，因为已主动校验
        return jsonify({"code": 400, "msg": "机构编码或名称已存在"}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e)}), 400

# @app.route('/api/office/upd', methods=['POST'])
# @token_required
# def update_office(current_user_id, current_user_admin):
#     data = request.get_json()
#     office_id = data.get('office_id')
#     if not office_id:
#         return jsonify({"code": 401, "msg": f"缺少 office_id"}), 200
#
#     office_code = data.get('office_code', '').strip()
#     office_name = data.get('office_name', '').strip()
#     if not office_code or not office_name:
#         return jsonify({"code": 401, "msg": f"机构编码和名称不能为空"}), 200
#
#     conn = get_db()
#     cursor = conn.cursor()
#     cursor.execute("SELECT office_id FROM sms_office WHERE office_id = ?", (office_id,))
#     if not cursor.fetchone():
#         return jsonify({"code": 401, "msg": f"机构不存在"}), 200
#
#     try:
#         cursor.execute(
#             "UPDATE sms_office SET office_code = ?, office_name = ? WHERE office_id = ?",
#             (office_code, office_name, office_id)
#         )
#         conn.commit()
#         cursor.close()
#         conn.close()
#         return jsonify({"code": 200, "msg": f"机构信息更新成功"}), 200
#
#     except sqlite3.IntegrityError as e:
#         return jsonify({"code": 401, "msg": f"机构编码或名称已存在"}), 200


@app.route('/api/user/change-password', methods=['POST'])
@token_required
def change_password(current_user_id, current_user_admin):
    data = request.get_json()
    userid = data.get('userid')
    password_old = data.get('password_old')
    password_new = data.get('password_new')

    if not userid or not password_old or not password_new:
        return jsonify({"code": 400, "msg": "userid, password_old, password_new 不能为空"}), 200

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, password FROM sms_user WHERE userid = ?", (userid,))
    user = cursor.fetchone()
    if not user:
        cursor.close()
        conn.close()
        return jsonify({"code": 404, "msg": "用户不存在"}), 200

    # 验证旧密码（存储的密文与传入的密文直接比较）
    if user['password'] != password_old:
        cursor.close()
        conn.close()
        return jsonify({"code": 401, "msg": "旧密码错误"}), 200

    # 更新为新密码（假设前端传入的 password_new 已经是密文，直接存储）
    cursor.execute("UPDATE sms_user SET password = ? WHERE id = ?", (password_new, user['id']))
    conn.commit()
    cursor.close()
    conn.close()

    # 可选：使该用户所有旧 token 失效（清除缓存中的 session_key）
    token_cache.delete(user['id'])

    return jsonify({"code": 200, "msg": "密码修改成功，请重新登录"}), 200

@app.route('/api/office/del/<int:office_id>', methods=['POST'])
@token_required
def delete_office(current_user_id, current_user_admin, office_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT office_id FROM sms_office WHERE office_id = ?", (office_id,))
    if not cursor.fetchone():
        return jsonify({"code": 401, "msg": f"机构不存在"}), 200

    # 将该机构下的用户的 office_id 置为 NULL
    cursor.execute("UPDATE sms_user SET office_id = NULL WHERE office_id = ?", (office_id,))
    # 删除机构
    cursor.execute("DELETE FROM sms_office WHERE office_id = ?", (office_id,))
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"code": 200, "msg": f"机构已删除，关联用户的机构信息已清空"}), 200



@app.route('/api/model-office/list', methods=['GET'])
@token_required
def model_office_list(current_user_id, current_user_admin):
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT admin_flg, office_id FROM sms_user WHERE id = ?", (current_user_id,))
    user_row = cursor.fetchone()
    if not user_row:
        return jsonify({"error": "用户不存在"}), 404

    is_super = user_row['admin_flg'] == 1 and not user_row['office_id']
    user_office_id = user_row['office_id']

    if is_super:
        # 超级管理员：所有机构
        cursor.execute("SELECT office_id, office_code, office_name FROM sms_office ORDER BY office_id")
        offices = cursor.fetchall()
        cursor.execute("SELECT office_id, model_name FROM model_office ORDER BY office_id, model_name")
        relations = cursor.fetchall()
        models_map = {}
        for row in relations:
            models_map.setdefault(row['office_id'], []).append(row['model_name'])
        result = []
        for off in offices:
            result.append({
                "office_id": off['office_id'],
                "office_code": off['office_code'],
                "office_name": off['office_name'],
                "model_names": models_map.get(off['office_id'], [])
            })
    else:
        # 非超级管理员：只能查看自己的机构（如果 office_id 为空，返回空）
        if not user_office_id:
            result = []
        else:
            cursor.execute("SELECT office_id, office_code, office_name FROM sms_office WHERE office_id = ?", (user_office_id,))
            office = cursor.fetchone()
            if office:
                cursor.execute("SELECT model_name FROM model_office WHERE office_id = ? ORDER BY model_name", (user_office_id,))
                rows = cursor.fetchall()
                model_names = [row['model_name'] for row in rows]
                result = [{
                    "office_id": office['office_id'],
                    "office_code": office['office_code'],
                    "office_name": office['office_name'],
                    "model_names": model_names
                }]
            else:
                result = []

    cursor.close()
    conn.close()
    return api_response(data=result)


@app.route('/api/model-office/new', methods=['POST'])
@token_required
def sync_model_office(current_user_id, current_user_admin):
    data = request.get_json()
    office_id = data.get('office_id')
    model_names = data.get('model_names', [])
    if not office_id:
        return jsonify({"code": 400, "msg": f"缺少 office_id"}), 200
    if not isinstance(model_names, list):
        return jsonify({"code": 400, "msg": f"model_names 必须是数组"}), 200

    conn = get_db()
    cursor = conn.cursor()
    # 检查机构是否存在
    cursor.execute("SELECT office_id FROM sms_office WHERE office_id = ?", (office_id,))
    if not cursor.fetchone():
        return jsonify({"code": 400, "msg": f"机构不存在"}), 200

    # 删除旧关联
    cursor.execute("DELETE FROM model_office WHERE office_id = ?", (office_id,))
    inserted = 0
    for mn in set(model_names):
        if mn and mn.strip():
            cursor.execute(
                "INSERT INTO model_office (office_id, model_name) VALUES (?, ?)",
                (office_id, mn.strip())
            )
            inserted += 1
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"code": 200, "msg": f"成功更新关联，共关联 {inserted} 个模型"}), 200


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=25125, debug=True)
